from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("docs/horse-racing-erd-explanation.xlsx")


TABLE_ROWS = [
    ["users", "id", "VARCHAR(64)", "PK", "Mã định danh duy nhất của tài khoản.", "u_owner"],
    ["users", "name", "VARCHAR(255)", "", "Tên hiển thị của người dùng.", "Sterling Stables"],
    ["users", "email", "VARCHAR(255)", "UNIQUE", "Email dùng để đăng nhập.", "owner@race.test"],
    ["users", "password", "VARCHAR(255)", "", "Mật khẩu đăng nhập local.", "owner123"],
    ["users", "role", "VARCHAR(32)", "", "Vai trò trong hệ thống: admin, owner, jockey, referee, spectator.", "owner"],
    ["users", "status", "VARCHAR(32)", "", "Trạng thái tài khoản: pending, active, rejected, suspended, locked.", "active"],

    ["jockeyProfiles", "id", "VARCHAR(64)", "PK", "Mã hồ sơ jockey.", "jp_001"],
    ["jockeyProfiles", "userId", "VARCHAR(64)", "FK -> users.id, UNIQUE", "Tài khoản jockey sở hữu profile.", "u_jockey"],
    ["jockeyProfiles", "bio", "TEXT", "", "Mô tả kinh nghiệm/chuyên môn.", "Experienced sprint jockey."],
    ["jockeyProfiles", "certificate", "TEXT", "", "Chứng chỉ hoặc giấy phép thi đấu.", "Class A Racing License"],
    ["jockeyProfiles", "competitionLevel", "VARCHAR(128)", "", "Cấp độ thi đấu.", "Elite"],
    ["jockeyProfiles", "weight", "NUMERIC(6,2)", "", "Cân nặng jockey, dùng cho tính handicap.", "54"],
    ["jockeyProfiles", "status", "VARCHAR(32)", "", "Trạng thái profile: draft, pending, published, rejected, archived.", "published"],
    ["jockeyProfiles", "updatedAt", "VARCHAR(64)", "", "Thời điểm cập nhật profile gần nhất.", "2026-06-01T08:00:00Z"],

    ["notifications", "id", "VARCHAR(64)", "PK", "Mã thông báo.", "n_001"],
    ["notifications", "userId", "VARCHAR(64)", "FK -> users.id", "Người nhận thông báo.", "u_owner"],
    ["notifications", "title", "VARCHAR(255)", "", "Tiêu đề thông báo.", "Race entry approved"],
    ["notifications", "message", "TEXT", "", "Nội dung thông báo.", "Midnight Storm has been approved."],
    ["notifications", "isRead", "BOOLEAN", "", "Đã đọc hay chưa.", "FALSE"],
    ["notifications", "createdAt", "VARCHAR(64)", "", "Thời điểm tạo thông báo.", "2026-06-01T08:00:00Z"],

    ["sessions", "token", "VARCHAR(128)", "PK", "Token đăng nhập.", "session_token"],
    ["sessions", "userId", "VARCHAR(64)", "FK -> users.id", "User đang sở hữu session.", "u_admin"],
    ["sessions", "createdAt", "VARCHAR(64)", "", "Thời điểm tạo session.", "2026-06-01T08:00:00Z"],

    ["tournaments", "id", "VARCHAR(64)", "PK", "Mã giải đấu.", "t_001"],
    ["tournaments", "name", "VARCHAR(255)", "", "Tên giải đấu.", "Summer Derby Classic"],
    ["tournaments", "status", "VARCHAR(64)", "", "Trạng thái tournament: registration, approvals, active, completed.", "approvals"],
    ["tournaments", "registrationWindow", "VARCHAR(128)", "", "Thời gian mở đăng ký.", "May 20 - May 31, 2026"],
    ["tournaments", "startDate", "VARCHAR(64)", "", "Ngày bắt đầu giải.", "June 10, 2026"],
    ["tournaments", "finalDate", "VARCHAR(64)", "", "Ngày kết thúc/final.", "June 30, 2026"],
    ["tournaments", "location", "VARCHAR(255)", "", "Địa điểm tổ chức.", "Churchill Downs"],
    ["tournaments", "prizePool", "NUMERIC(14,2)", "", "Tổng giải thưởng.", "750000"],

    ["races", "id", "VARCHAR(64)", "PK", "Mã cuộc đua.", "r_001"],
    ["races", "tournamentId", "VARCHAR(64)", "FK -> tournaments.id", "Race thuộc tournament nào.", "t_001"],
    ["races", "raceNumber", "VARCHAR(64)", "", "Số race trong tournament.", "R1"],
    ["races", "name", "VARCHAR(255)", "", "Tên race.", "Summer Derby Qualifier R1"],
    ["races", "round", "VARCHAR(64)", "", "Vòng thi đấu.", "Qualifier"],
    ["races", "date", "VARCHAR(64)", "", "Ngày đua.", "June 10, 2026"],
    ["races", "time", "VARCHAR(32)", "", "Giờ đua.", "16:30"],
    ["races", "venue", "VARCHAR(255)", "", "Sân/địa điểm race.", "Churchill Downs"],
    ["races", "distance", "VARCHAR(64)", "", "Cự ly đua.", "1400m"],
    ["races", "surface", "VARCHAR(64)", "", "Mặt sân.", "Turf"],
    ["races", "raceClass", "VARCHAR(128)", "", "Hạng race.", "Class A"],
    ["races", "handicapMin", "NUMERIC(6,2)", "", "Mức handicap tối thiểu.", "0"],
    ["races", "handicapMax", "NUMERIC(6,2)", "", "Mức handicap tối đa.", "5"],
    ["races", "status", "VARCHAR(64)", "", "Trạng thái race: draft, registration-open, published, in-progress, finished, completed.", "published"],
    ["races", "resultStatus", "VARCHAR(32)", "", "Trạng thái kết quả: draft, submitted, approved, rejected.", "draft"],
    ["races", "awardsPublished", "BOOLEAN", "", "Đã công bố awards hay chưa.", "FALSE"],
    ["races", "createdBy", "VARCHAR(64)", "FK -> users.id", "Admin tạo race.", "u_admin"],

    ["horses", "id", "VARCHAR(64)", "PK", "Mã ngựa.", "h_001"],
    ["horses", "ownerUserId", "VARCHAR(64)", "FK -> users.id", "Owner sở hữu ngựa.", "u_owner"],
    ["horses", "selectedJockeyUserId", "VARCHAR(64)", "FK -> users.id", "Jockey được owner chọn gần nhất.", "u_jockey"],
    ["horses", "name", "VARCHAR(255)", "", "Tên ngựa.", "Midnight Storm"],
    ["horses", "breed", "VARCHAR(128)", "", "Giống ngựa.", "Thoroughbred"],
    ["horses", "species", "VARCHAR(128)", "", "Loài.", "Equus ferus caballus"],
    ["horses", "age", "INTEGER", "", "Tuổi ngựa.", "4"],
    ["horses", "sex", "VARCHAR(64)", "", "Giới tính.", "Stallion"],
    ["horses", "color", "VARCHAR(128)", "", "Màu lông.", "Black"],
    ["horses", "weightKg", "NUMERIC(7,2)", "", "Cân nặng ngựa theo kg.", "485"],
    ["horses", "heightCm", "NUMERIC(7,2)", "", "Chiều cao ngựa theo cm.", "164"],
    ["horses", "baseHandicap", "NUMERIC(6,2)", "", "Handicap nền trước khi race.", "5"],
    ["horses", "speedRating", "NUMERIC(6,2)", "", "Chỉ số tốc độ.", "88"],
    ["horses", "staminaRating", "NUMERIC(6,2)", "", "Chỉ số sức bền.", "82"],
    ["horses", "formRating", "NUMERIC(6,2)", "", "Phong độ gần đây.", "86"],
    ["horses", "healthRating", "NUMERIC(6,2)", "", "Chỉ số sức khỏe.", "92"],
    ["horses", "overallRating", "NUMERIC(6,2)", "", "Rating tổng hợp.", "86.80"],
    ["horses", "status", "VARCHAR(32)", "", "Trạng thái hồ sơ: pending, approved, rejected, retired.", "approved"],
    ["horses", "jockeyConfirmation", "VARCHAR(64)", "", "Trạng thái xác nhận jockey.", "confirmed"],

    ["jockeyTournamentRegistrations", "id", "VARCHAR(64)", "PK", "Mã đăng ký jockey vào tournament.", "jtr_001"],
    ["jockeyTournamentRegistrations", "tournamentId", "VARCHAR(64)", "FK -> tournaments.id", "Tournament jockey muốn tham gia.", "t_001"],
    ["jockeyTournamentRegistrations", "jockeyUserId", "VARCHAR(64)", "FK -> users.id", "Tài khoản jockey đăng ký.", "u_jockey"],
    ["jockeyTournamentRegistrations", "status", "VARCHAR(32)", "", "Trạng thái duyệt: pending, approved, rejected.", "approved"],
    ["jockeyTournamentRegistrations", "createdAt", "VARCHAR(64)", "", "Thời điểm jockey gửi request.", "2026-05-20T08:00:00Z"],
    ["jockeyTournamentRegistrations", "reviewedAt", "VARCHAR(64)", "", "Thời điểm admin duyệt/từ chối.", "2026-05-20T09:00:00Z"],

    ["jockeyInvitations", "id", "VARCHAR(64)", "PK", "Mã lời mời jockey.", "ji_001"],
    ["jockeyInvitations", "horseId", "VARCHAR(64)", "FK -> horses.id", "Ngựa được đăng ký.", "h_001"],
    ["jockeyInvitations", "ownerUserId", "VARCHAR(64)", "FK -> users.id", "Owner gửi lời mời.", "u_owner"],
    ["jockeyInvitations", "jockeyUserId", "VARCHAR(64)", "FK -> users.id", "Jockey được mời.", "u_jockey"],
    ["jockeyInvitations", "tournamentId", "VARCHAR(64)", "FK -> tournaments.id", "Tournament liên quan.", "t_001"],
    ["jockeyInvitations", "raceId", "VARCHAR(64)", "FK -> races.id", "Race cụ thể.", "r_001"],
    ["jockeyInvitations", "status", "VARCHAR(32)", "", "Phản hồi jockey: pending, accepted, rejected, cancelled.", "accepted"],
    ["jockeyInvitations", "adminStatus", "VARCHAR(32)", "", "Trạng thái admin duyệt pairing.", "approved"],
    ["jockeyInvitations", "createdAt", "VARCHAR(64)", "", "Thời điểm tạo lời mời.", "2026-05-30T06:36:52Z"],
    ["jockeyInvitations", "respondedAt", "VARCHAR(64)", "", "Thời điểm jockey phản hồi.", "2026-05-30T06:37:25Z"],

    ["raceRefereeAssignments", "id", "VARCHAR(64)", "PK", "Mã phân công referee.", "rra_r_001_u_referee"],
    ["raceRefereeAssignments", "raceId", "VARCHAR(64)", "FK -> races.id", "Race được phân công.", "r_001"],
    ["raceRefereeAssignments", "refereeUserId", "VARCHAR(64)", "FK -> users.id", "Referee được phân công.", "u_referee"],
    ["raceRefereeAssignments", "assignedBy", "VARCHAR(64)", "FK -> users.id", "Admin phân công.", "u_admin"],
    ["raceRefereeAssignments", "status", "VARCHAR(32)", "", "Trạng thái phân công: assigned, confirmed, declined, removed.", "assigned"],
    ["raceRefereeAssignments", "assignedAt", "VARCHAR(64)", "", "Thời điểm phân công.", "2026-06-01T08:00:00Z"],

    ["raceEntries", "id", "VARCHAR(64)", "PK", "Mã entry tham gia race.", "re_001"],
    ["raceEntries", "raceId", "VARCHAR(64)", "FK -> races.id", "Race mà ngựa tham gia.", "r_001"],
    ["raceEntries", "horseId", "VARCHAR(64)", "FK -> horses.id", "Ngựa tham gia.", "h_001"],
    ["raceEntries", "jockeyUserId", "VARCHAR(64)", "FK -> users.id", "Jockey cưỡi ngựa trong race.", "u_jockey"],
    ["raceEntries", "invitationId", "VARCHAR(64)", "FK -> jockeyInvitations.id", "Lời mời tạo ra entry.", "ji_001"],
    ["raceEntries", "status", "VARCHAR(32)", "", "Trạng thái entry: pending-approval, approved, rejected.", "approved"],
    ["raceEntries", "lane", "INTEGER", "", "Gate/line xuất phát.", "1"],
    ["raceEntries", "handicap", "NUMERIC(6,2)", "", "Handicap áp dụng trong race.", "5.5"],
    ["raceEntries", "ratingSnapshot", "NUMERIC(6,2)", "", "Rating tại thời điểm chuẩn bị race.", "86.8"],
    ["raceEntries", "ownerConfirmed", "BOOLEAN", "", "Owner đã xác nhận tham gia.", "TRUE"],
    ["raceEntries", "jockeyConfirmed", "BOOLEAN", "", "Jockey đã xác nhận tham gia.", "TRUE"],
    ["raceEntries", "preRaceStatus", "VARCHAR(32)", "", "Tình trạng trước race: pending, ready, absent.", "ready"],
    ["raceEntries", "disqualified", "BOOLEAN", "", "Bị loại khỏi race hay không.", "FALSE"],
    ["raceEntries", "resultStatus", "VARCHAR(32)", "", "Trạng thái kết quả: draft, submitted, official.", "official"],
    ["raceEntries", "position", "INTEGER", "", "Thứ hạng về đích.", "1"],
    ["raceEntries", "finishTime", "VARCHAR(64)", "", "Thời gian hoàn thành.", "01:18.52"],
    ["raceEntries", "notes", "TEXT", "", "Ghi chú kết quả.", "R1 official result - 10 pts"],
    ["raceEntries", "violationNotes", "TEXT", "", "Ghi chú vi phạm nếu có.", ""],

    ["refereeReports", "id", "VARCHAR(64)", "PK", "Mã báo cáo referee.", "rr_001"],
    ["refereeReports", "raceId", "VARCHAR(64)", "FK -> races.id", "Race có báo cáo.", "r_001"],
    ["refereeReports", "raceEntryId", "VARCHAR(64)", "FK -> raceEntries.id", "Entry liên quan đến báo cáo.", "re_001"],
    ["refereeReports", "refereeUserId", "VARCHAR(64)", "FK -> users.id", "Referee tạo báo cáo.", "u_referee"],
    ["refereeReports", "reportType", "VARCHAR(64)", "", "Loại báo cáo: incident, violation, inspection.", "violation"],
    ["refereeReports", "description", "TEXT", "", "Mô tả sự cố/vi phạm.", "False start detected."],
    ["refereeReports", "violation", "TEXT", "", "Nội dung vi phạm cụ thể.", "False start"],
    ["refereeReports", "status", "VARCHAR(32)", "", "Trạng thái report: draft, submitted, reviewed, dismissed.", "submitted"],
    ["refereeReports", "createdAt", "VARCHAR(64)", "", "Thời điểm tạo report.", "2026-06-10T10:00:00Z"],
    ["refereeReports", "reviewedAt", "VARCHAR(64)", "", "Thời điểm admin review.", "2026-06-10T11:00:00Z"],
]


RELATIONSHIP_ROWS = [
    ["users", "sessions", "users.id -> sessions.userId", "1 user có nhiều session", "Lưu phiên đăng nhập."],
    ["users", "notifications", "users.id -> notifications.userId", "1 user nhận nhiều notification", "Thông báo trạng thái nghiệp vụ."],
    ["users", "jockeyProfiles", "users.id -> jockeyProfiles.userId", "1 jockey có 0 hoặc 1 profile", "Hồ sơ công khai của jockey."],
    ["users", "horses", "users.id -> horses.ownerUserId", "1 owner sở hữu nhiều horse", "Owner quản lý ngựa."],
    ["tournaments", "races", "tournaments.id -> races.tournamentId", "1 tournament có nhiều race", "Admin tạo các R1-R10 trong tournament."],
    ["tournaments", "jockeyTournamentRegistrations", "tournaments.id -> jockeyTournamentRegistrations.tournamentId", "1 tournament có nhiều request jockey", "Jockey xin tham gia tournament."],
    ["users", "jockeyTournamentRegistrations", "users.id -> jockeyTournamentRegistrations.jockeyUserId", "1 jockey có nhiều tournament request", "Admin duyệt jockey theo tournament."],
    ["horses", "jockeyInvitations", "horses.id -> jockeyInvitations.horseId", "1 horse có nhiều invitation theo race", "Owner mời jockey cho horse."],
    ["users", "jockeyInvitations", "users.id -> ownerUserId/jockeyUserId", "Owner gửi, Jockey nhận", "Luồng accept/reject trước khi admin duyệt."],
    ["races", "jockeyInvitations", "races.id -> jockeyInvitations.raceId", "1 race có nhiều invitation", "Owner chọn jockey cho race cụ thể."],
    ["jockeyInvitations", "raceEntries", "jockeyInvitations.id -> raceEntries.invitationId", "1 accepted invitation tạo 0 hoặc 1 race entry", "Admin approve pairing thì tạo entry."],
    ["races", "raceEntries", "races.id -> raceEntries.raceId", "1 race có nhiều entry", "Danh sách ngựa/jockey trong race."],
    ["horses", "raceEntries", "horses.id -> raceEntries.horseId", "1 horse có nhiều race entry", "Ngựa có thể chạy nhiều race."],
    ["users", "raceEntries", "users.id -> raceEntries.jockeyUserId", "1 jockey có nhiều race entry", "Jockey cưỡi ngựa trong từng race."],
    ["races", "raceRefereeAssignments", "races.id -> raceRefereeAssignments.raceId", "1 race có nhiều referee", "Admin phân công referee."],
    ["users", "raceRefereeAssignments", "users.id -> raceRefereeAssignments.refereeUserId", "1 referee được phân công nhiều race", "Referee chỉ thấy race được giao."],
    ["races", "refereeReports", "races.id -> refereeReports.raceId", "1 race có nhiều report", "Ghi nhận incident/violation cấp race."],
    ["raceEntries", "refereeReports", "raceEntries.id -> refereeReports.raceEntryId", "1 entry có nhiều report", "Report cho ngựa/jockey cụ thể."],
    ["users", "refereeReports", "users.id -> refereeReports.refereeUserId", "1 referee tạo nhiều report", "Truy vết người lập report."],
]


FLOW_ROWS = [
    [1, "Admin tạo tournament", "tournaments", "Tạo giải đấu trước khi có race."],
    [2, "Admin tạo race", "races", "Mỗi tournament có nhiều race, ví dụ R1-R10."],
    [3, "Admin phân công referee", "raceRefereeAssignments", "Race có thể có nhiều referee."],
    [4, "Owner tạo hồ sơ horse", "horses", "Horse chờ admin duyệt trước khi đăng ký race."],
    [5, "Jockey tạo profile", "jockeyProfiles", "Profile published thì các role có thể xem."],
    [6, "Jockey đăng ký tournament", "jockeyTournamentRegistrations", "Admin approve thì owner mới chọn được jockey trong tournament đó."],
    [7, "Owner chọn horse + race + jockey", "jockeyInvitations", "Tạo lời mời để jockey xác nhận."],
    [8, "Jockey accept invitation", "jockeyInvitations", "Nếu accept thì chuyển qua admin duyệt pairing."],
    [9, "Admin approve pairing", "raceEntries", "Tạo race entry chính thức."],
    [10, "Đóng registration", "raceEntries", "Hệ thống gán gate, handicap, rating snapshot."],
    [11, "Referee kiểm tra participant", "raceEntries", "Mark ready/absent trước khi start race."],
    [12, "Referee start race", "races", "Race chuyển sang in-progress."],
    [13, "Referee ghi kết quả", "raceEntries", "Nhập position, finishTime, notes."],
    [14, "Referee report violation", "refereeReports", "Lưu incident/violation nếu có."],
    [15, "Admin confirm result", "races / raceEntries", "Kết quả thành official và hiển thị cho các role."],
]


def style_sheet(ws, title, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="0B223D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = ws.dimensions


def add_title(ws, text):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    cell = ws.cell(1, 1)
    cell.value = text
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="12395A")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "ERD Tables"
    ws.append(["Table", "Field", "Suggested Data Type", "Key", "Meaning / Business Purpose", "Example"])
    for row in TABLE_ROWS:
        ws.append(row)
    style_sheet(ws, "ERD Tables", [28, 26, 22, 22, 62, 30])
    add_title(ws, "ERD Table Dictionary - Horse Racing Tournament System")

    ws_rel = wb.create_sheet("Relationships")
    ws_rel.append(["From Table", "To Table", "Join / FK", "Cardinality", "Business Meaning"])
    for row in RELATIONSHIP_ROWS:
        ws_rel.append(row)
    style_sheet(ws_rel, "Relationships", [28, 32, 42, 34, 58])
    add_title(ws_rel, "ERD Relationships")

    ws_flow = wb.create_sheet("Business Flow")
    ws_flow.append(["Step", "Business Action", "Main Table", "Explanation"])
    for row in FLOW_ROWS:
        ws_flow.append(row)
    style_sheet(ws_flow, "Business Flow", [10, 36, 32, 72])
    add_title(ws_flow, "End-to-End Business Flow Mapped To ERD")

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Component", "Explanation"])
    summary_rows = [
        ["Core idea", "raceEntries là bảng trung tâm, nối Race + Horse + Jockey và lưu gate, handicap, result."],
        ["Identity", "users lưu tất cả account. Role được phân biệt bằng users.role."],
        ["Approval", "Owner/Jockey/Referee account, horse, jockey tournament registration và pairing đều cần admin duyệt tùy rule."],
        ["Race setup", "Admin tạo tournament, race và raceRefereeAssignments."],
        ["Race operation", "Referee check participant, start race, record result và tạo refereeReports nếu có vi phạm."],
        ["Result", "Admin confirm result, sau đó kết quả official mới dùng cho ranking/awards."],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    style_sheet(ws_summary, "Summary", [24, 90])
    add_title(ws_summary, "Presentation Summary")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
